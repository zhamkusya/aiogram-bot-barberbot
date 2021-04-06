from config import TOKEN
import pandas as pd
import logging
from aiogram import Bot, Dispatcher, executor, types
import knopkaa as kb
from aiogram.dispatcher.filters import Text
from openpyxl import load_workbook
wb = load_workbook(filename="botbarber.xlsx")
sheet = wb.active

bot = Bot(token=TOKEN)
dp = Dispatcher(bot)
wb = load_workbook(filename="botbarber.xlsx")
sheet = wb.active
barberslist=[]
service=[]
for row in sheet.rows:
    a=(str(row[0].value))
    barberslist.append(a)
print(barberslist)
for row in sheet.rows:
    a=(str(row[1].value))
    service.append(a)
print(service)


@dp.message_handler(commands=['start'])
async def process_hello(message: types.Message):
      await bot.send_message(message.from_user.id, "Добро пожаловать в «Барбершоп Uncle Chill». Барбершоп «Дядя Chill» — это место, где можно выпить чашечку кофе или угоститься другим любимым напитком, поиграть в бильярд, сделать стрижку или тату, провести мужскую косметологию, маникюр и педикюр, оставаясь верным традициям мужского этикета.С помощью нашего бота, вы сможете записаться онлайн на любой вид наших услуг, ознакомиться с ценами и барным меню, а так же задать любой интересующий Вас вопрос",reply_markup=kb.menu)

@dp.message_handler(Text(equals="🧭 Онлайн-запись"))
async def with_puree(message: types.Message):
    await bot.send_message(message.from_user.id, 'Выберите категорию услуг из списка ниже ⬇️', reply_markup=kb.onlzap)

@dp.message_handler(Text(equals="💇‍♂️ Мужские стрижки"))
async def barbers(message: types.Message):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add("🔄 Пропустить")
    for i in barberslist:
        keyboard.add(i)
    keyboard.add("⬅️ Назад")
    await bot.send_message(message.from_user.id,"Выберите мастера из списка ниже ⬇️",  reply_markup=keyboard)
@dp.message_handler(Text(equals="🔄 Пропустить"))
async def propust(message: types.Message):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    for i in service:
        keyboard.add(i)
    keyboard.add("⬅️ Назад")
    await bot.send_message(message.from_user.id,"Выберите услуги из списка ниже ⬇️",  reply_markup=keyboard)

@dp.message_handler(Text(equals="Стрижка"))
async def propust1(message: types.Message):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add("⬅️ Оформить бронирование")
    for i in service:
        if i=="Стрижка":
            keyboard.add(i+"✅")
        else:
            keyboard.add(i)
    keyboard.add("⬅️ Назад")
    await bot.send_message(message.from_user.id,"Выберите услуги из списка ниже ⬇️",  reply_markup=keyboard)
@dp.message_handler(Text(equals="Детская стрижка до 9 лет"))
async def propust(message: types.Message):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add("⬅️ Оформить бронирование")
    for i in service:
        if i=="Детская стрижка до 9 лет":
            keyboard.add(i+"✅")
        else:
            keyboard.add(i)
    keyboard.add("⬅️ Назад")
    await bot.send_message(message.from_user.id, "Выберите услуги из списка ниже ⬇️", reply_markup=keyboard)
@dp.message_handler(Text(equals="Королевское бритье"))
async def propust(message: types.Message):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add("⬅️ Оформить бронирование")
    for i in service:
        if i=="Королевское бритье":
            keyboard.add(i+"✅")
        else:
            keyboard.add(i)
    keyboard.add("⬅️ Назад")
    await bot.send_message(message.from_user.id, "Выберите услуги из списка ниже ⬇️", reply_markup=keyboard)
@dp.message_handler(Text(equals="Бритье головы опасной бритвой"))
async def propust(message: types.Message):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add("⬅️ Оформить бронирование")
    for i in service:
        if i=="Бритье головы опасной бритвой":
            keyboard.add(i+"✅")
        else:
            keyboard.add(i)
    keyboard.add("⬅️ Назад")
    await bot.send_message(message.from_user.id, "Выберите услуги из списка ниже ⬇️", reply_markup=keyboard)
@dp.message_handler(Text(equals="Моделирование усов и бороды"))
async def propust(message: types.Message):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add("⬅️ Оформить бронирование")
    for i in service:
        if i=="Моделирование усов и бороды":
            keyboard.add(i+"✅")
        else:
            keyboard.add(i)
    keyboard.add("⬅️ Назад")
    await bot.send_message(message.from_user.id, "Выберите услуги из списка ниже ⬇️", reply_markup=keyboard)
@dp.message_handler(Text(equals="Мойка массаж головы"))
async def propust(message: types.Message):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add("⬅️ Оформить бронирование")
    for i in service:
        if i=="Мойка массаж головы":
            keyboard.add(i+"✅")
        else:
            keyboard.add(i)
    keyboard.add("⬅️ Назад")
    await bot.send_message(message.from_user.id, "Выберите услуги из списка ниже ⬇️", reply_markup=keyboard)
@dp.message_handler(Text(equals="Стрижка отец и сын"))
async def propust(message: types.Message):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add("⬅️ Оформить бронирование")
    for i in service:
        if i=="Стрижка отец и сын":
            keyboard.add(i+"✅")
        else:
            keyboard.add(i)
    keyboard.add("⬅️ Назад")
    await bot.send_message(message.from_user.id, "Выберите услуги из списка ниже ⬇️", reply_markup=keyboard)
@dp.message_handler(Text(equals="⬅️ Назад"))
async def propust2(message: types.Message):
    keyboard = types.ReplyKeyboardMarkup(resize_keyboard=True)
    keyboard.add("🔄 Пропустить")
    for i in barberslist:
        keyboard.add(i)
    keyboard.add("⬅️ Назад")
    await bot.send_message(message.from_user.id, "Выберите мастера из списка ниже ⬇️", reply_markup=keyboard)

@dp.message_handler(Text(equals="⬅️ Главное меню"))
async def process_hello(message: types.Message):
      await bot.send_message(message.from_user.id, "Добро пожаловать в «Барбершоп Uncle Chill». Барбершоп «Дядя Chill» — это место, где можно выпить чашечку кофе или угоститься другим любимым напитком, поиграть в бильярд, сделать стрижку или тату, провести мужскую косметологию, маникюр и педикюр, оставаясь верным традициям мужского этикета.С помощью нашего бота, вы сможете записаться онлайн на любой вид наших услуг, ознакомиться с ценами и барным меню, а так же задать любой интересующий Вас вопрос",reply_markup=kb.menu)

@dp.message_handler(Text(equals="☎️ Контакты"))
async def with_puree(message: types.Message):
    await bot.send_message(message.from_user.id, "📍 Адрес: г.Ташкент, ул. Матбуотчилар, 17"
"☎️ Номер телефона: +998 97 776-66-92"
"🌎 Сайт: http://unchlechill.uz"
"📌 Facebook: https://www.facebook.com/barbershop.unclechill/"
"📌 Instagram: https://www.instagram.com/barbershop_unclechill/"
"🤖 Telegram-бот: https://t.me/unclechill_bot️")
@dp.message_handler(Text(equals="❓ Задать вопрос"))
async def with_puree(message: types.Message):
    await bot.send_message(message.from_user.id, 'Введите ваш вопрос и на него ответят в ближайшее время наши сотрудники.️', reply_markup=kb.questkn)



if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)
