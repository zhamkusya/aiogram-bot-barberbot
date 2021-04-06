from aiogram.types import ReplyKeyboardRemove, \
    ReplyKeyboardMarkup, KeyboardButton, \
    InlineKeyboardMarkup, InlineKeyboardButton
from openpyxl import load_workbook
from config import TOKEN
import pandas as pd
import logging
from aiogram import Bot, Dispatcher, executor, types
import knopkaa as kb
from aiogram.dispatcher.filters import Text
from openpyxl import load_workbook
wb = load_workbook(filename="botbarber.xlsx")
sheet = wb.active
barberslist=[]
for x in sheet.rows:
    barberslist.append(x)


menu = ReplyKeyboardMarkup(
    keyboard=[
        [
            KeyboardButton(text="🧭 Онлайн-запись"),
        ],
        [
            KeyboardButton(text="☎️ Контакты"),
            KeyboardButton(text="❓ Задать вопрос")
        ],

[
            KeyboardButton(text="🖼 Галерея"),
            KeyboardButton(text="💰 Цены на услуги")
        ],
        [ KeyboardButton(text="⚙️ Настройки")]],
    resize_keyboard=True
)
onlzap = ReplyKeyboardMarkup(
    keyboard=[
        [
            KeyboardButton(text="💇‍♂️ Мужские стрижки"),
        ],
        [
            KeyboardButton(text="⬅️ Главное меню")
        ]],
    resize_keyboard=True
)
questkn=ReplyKeyboardMarkup(
    keyboard=[

        [
            KeyboardButton(text="⬅️ Главное меню")
        ]],
    resize_keyboard=True
)

